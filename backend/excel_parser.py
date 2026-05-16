from __future__ import annotations

import os
import re
from functools import cached_property
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

import models

# ─── Constants ───────────────────────────────────────────────────────────────

_UNKNOWN = "Не указано"

# Filename convention: 028_Фамилия_Имя_Отчество_141а_ПИо_...
_IDX_CARD   = 0
_IDX_LAST   = 1
_IDX_FIRST  = 2
_IDX_MIDDLE = 3
_IDX_GROUP  = 4
_IDX_CODE   = 5

# Cell addresses in the "Титул" sheet
_CELL_FULL_NAME      = "H27"
_CELL_PROGRAM        = "D29"
_CELL_SPECIALIZATION = "D30"


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _to_int(value) -> int:
    """Convert a cell value (string, float, int, or None) to int, defaulting to 0."""
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


# ─── Parser ──────────────────────────────────────────────────────────────────

class ExcelParser:
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)

    # ── Internal helpers ────────────────────────────────────────────────────

    def _clean(self, value) -> str:
        if value is None:
            return ""
        return str(value).replace("\r", "").strip()

    @cached_property
    def _filename_parts(self) -> list[str]:
        stem = os.path.splitext(os.path.basename(self.file_path))[0]
        return stem.split("_")

    def _student_card(self) -> str:
        parts = self._filename_parts
        if parts and parts[_IDX_CARD].isdigit():
            return parts[_IDX_CARD]
        return str(abs(hash(os.path.basename(self.file_path))))[:8]

    def _group_name(self) -> str:
        parts = self._filename_parts
        if len(parts) > _IDX_CODE:
            return f"{parts[_IDX_GROUP]}-{parts[_IDX_CODE]}"
        if len(parts) > _IDX_GROUP:
            return parts[_IDX_GROUP]
        return _UNKNOWN

    def _full_name_from_filename(self) -> str:
        parts = self._filename_parts
        if len(parts) > _IDX_MIDDLE:
            return f"{parts[_IDX_LAST]} {parts[_IDX_FIRST]} {parts[_IDX_MIDDLE]}"
        return "Неизвестный студент"

    def _course_from_filename(self) -> Optional[int]:
        parts = self._filename_parts
        group = parts[_IDX_GROUP] if len(parts) > _IDX_GROUP else ""
        m = re.match(r"(\d)", group)
        return int(m.group(1)) if m else None

    # ── Public parse methods ─────────────────────────────────────────────────

    def parse_student_info(self) -> dict:
        sheet = self.workbook["Титул"] if "Титул" in self.workbook.sheetnames else self.workbook.active

        full_name = self._clean(sheet[_CELL_FULL_NAME].value)
        if full_name:
            lines = [ln.strip() for ln in full_name.split("\n") if ln.strip()]
            full_name = lines[-1] if lines else full_name
        if not full_name:
            full_name = self._full_name_from_filename()

        program_name = self._clean(sheet[_CELL_PROGRAM].value)
        if program_name and " " in program_name:
            program_name = program_name.split(" ", 1)[1].strip()

        specialization = self._clean(sheet[_CELL_SPECIALIZATION].value)

        return {
            "full_name": full_name,
            "student_card_number": self._student_card(),
            "group_name": self._group_name(),
            "course": self._course_from_filename(),
            "program_name": program_name or _UNKNOWN,
            "specialization": specialization or _UNKNOWN,
        }

    def parse_attested_items(self) -> list[dict]:
        if "Переаттестация" not in self.workbook.sheetnames:
            return []

        sheet = self.workbook["Переаттестация"]
        seen: dict[str, dict] = {}
        current_name = ""

        for row in range(4, sheet.max_row + 1):
            name_val = sheet.cell(row=row, column=2).value
            if name_val:
                current_name = self._clean(name_val)
            if not current_name:
                continue

            control_form     = sheet.cell(row=row, column=9).value
            attestation_type = sheet.cell(row=row, column=10).value
            result_val       = sheet.cell(row=row, column=11).value

            if not (attestation_type or result_val or control_form):
                continue

            seen[current_name] = {
                "discipline_name": current_name,
                "credits": _to_int(sheet.cell(row=row, column=7).value),
                "control_form": self._clean(control_form) or _UNKNOWN,
                "result": self._clean(result_val) or _UNKNOWN,
                "attestation_type": self._clean(attestation_type) or _UNKNOWN,
            }

        return list(seen.values())

    def parse_plan_items(self) -> list[dict]:
        if "План" not in self.workbook.sheetnames:
            return []

        sheet = self.workbook["План"]
        seen: dict[str, dict] = {}

        for row in range(6, sheet.max_row + 1):
            name = self._clean(sheet.cell(row=row, column=3).value)
            if not name:
                continue

            lowered = name.lower()
            if lowered.startswith("модуль") or "блок" in lowered or "часть" in lowered:
                continue

            exam             = sheet.cell(row=row, column=4).value
            credit           = sheet.cell(row=row, column=5).value
            credit_with_mark = sheet.cell(row=row, column=6).value
            course_work      = sheet.cell(row=row, column=7).value

            if exam:             control_form = "Экзамен"
            elif credit:         control_form = "Зачет"
            elif credit_with_mark: control_form = "Зачет с оценкой"
            elif course_work:    control_form = "Курсовая работа"
            else:                control_form = _UNKNOWN

            seen[name] = {
                "discipline_name": name,
                "credits": _to_int(sheet.cell(row=row, column=8).value),
                "hours": _to_int(sheet.cell(row=row, column=9).value),
                "control_form": control_form,
            }

        return list(seen.values())

    def parse_debts(
        self,
        plan_items: Optional[list[dict]] = None,
        attested_items: Optional[list[dict]] = None,
    ) -> list[dict]:
        if plan_items is None:
            plan_items = self.parse_plan_items()
        if attested_items is None:
            attested_items = self.parse_attested_items()

        attested_names = {
            item["discipline_name"].lower()
            for item in attested_items
            if item.get("discipline_name")
        }

        return [
            {
                "name": item["discipline_name"],
                "remaining_credits": item["credits"],
                "remaining_hours": item["hours"],
                "control_form": item["control_form"],
                "deadline": None,
            }
            for item in plan_items
            if item["discipline_name"].lower() not in attested_names
        ]


# ─── Loader ──────────────────────────────────────────────────────────────────

class ExcelDataLoader:
    def __init__(self, db_session: Session) -> None:
        self.db = db_session

    def save_to_database(self, file_path: str, student_id: Optional[int] = None) -> dict:
        try:
            parser = ExcelParser(file_path)
            student_info = parser.parse_student_info()
            card = student_info["student_card_number"]

            db_student: Optional[models.Student] = None
            if student_id is not None:
                db_student = self.db.query(models.Student).filter(models.Student.id == student_id).first()
            if db_student is None:
                db_student = self.db.query(models.Student).filter(models.Student.student_card_number == card).first()

            if db_student is not None:
                for field, value in student_info.items():
                    setattr(db_student, field, value)
                db_student.status = "active"
            else:
                db_student = models.Student(**student_info, status="active")
                self.db.add(db_student)
                self.db.flush()

            sid = db_student.id

            # Replace all dependent data atomically
            self.db.query(models.IndividualPlanItem).filter(models.IndividualPlanItem.student_id == sid).delete(synchronize_session=False)
            self.db.query(models.AttestationItem).filter(models.AttestationItem.student_id == sid).delete(synchronize_session=False)
            self.db.query(models.Debt).filter(models.Debt.student_id == sid).delete(synchronize_session=False)

            attested_items = parser.parse_attested_items()
            plan_items     = parser.parse_plan_items()
            debts          = parser.parse_debts(plan_items=plan_items, attested_items=attested_items)

            for item in attested_items:
                self.db.add(models.AttestationItem(
                    student_id=sid,
                    discipline_name=item["discipline_name"],
                    credits=item["credits"],
                    control_form=item["control_form"],
                    result=item["result"],
                    attestation_type=item["attestation_type"],
                ))

            created_debts: list[models.Debt] = []
            for item in debts:
                debt = models.Debt(
                    student_id=sid,
                    discipline_name=item["name"],
                    hours_remaining=item["remaining_hours"],
                    credits_remaining=item["remaining_credits"],
                    reassessment_deadline=item["deadline"],
                    status="active",
                )
                self.db.add(debt)
                self.db.flush()
                created_debts.append(debt)

            for idx, item in enumerate(debts):
                self.db.add(models.IndividualPlanItem(
                    student_id=sid,
                    debt_id=created_debts[idx].id,
                    discipline_name=item["name"],
                    hours=item["remaining_hours"],
                    credits=item["remaining_credits"],
                    control_form=item["control_form"],
                    deadline=item["deadline"],
                ))

            self.db.commit()

            return {
                "success": True,
                "student_id": sid,
                "disciplines_count": len(attested_items),
                "debts_count": len(debts),
                "plan_items_count": len(debts),
            }

        except Exception as exc:
            self.db.rollback()
            return {"success": False, "student_id": None, "error": str(exc)}
